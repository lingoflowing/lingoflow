#!/usr/bin/env python3
"""
LingoFlow v2 data builder

Input:
  CardMaster600 / ImageMaster600 style xlsx

Output:
  data/cards.json
  data/images.json
  data/chapters.json
  data/playlists.json

Usage:
  python tools/build_data_from_xlsx.py ImageMaster600_v36_Semantic_Rebuild_02.xlsx
"""
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl is required: pip install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data"


def text(value):
    return "" if value is None else str(value).strip()


def card_no3(value):
    match = re.search(r"\d+", text(value))
    return f"{int(match.group(0)):03d}" if match else "000"


def main(path):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["CardMaster600"] if "CardMaster600" in workbook.sheetnames else workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    def get(row, name):
        if name not in headers:
            return ""
        value = row[headers.index(name)]
        return text(value)

    cards = []
    images = []
    chapters = {}
    playlists = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        no = card_no3(get(row, "CardNo"))
        card_id = f"card_{no}"
        image_file = get(row, "ImageFile") or f"{card_id}.webp"
        image_path = image_file if image_file.startswith("images/") else f"images/{image_file}"
        if not re.search(r"\.(png|jpg|jpeg|webp|gif|avif)$", image_path, re.I):
            image_path += ".webp"

        chapter_no = get(row, "ChapterNo")
        playlist_no = get(row, "PlaylistNo")

        card = {
            "id": card_id,
            "cardNo": no,
            "chapterNo": chapter_no,
            "chapterTitle": get(row, "ChapterTitle"),
            "playlistNo": playlist_no,
            "playlistTitle": get(row, "PlaylistTitle"),
            "wordZh": get(row, "WordZh"),
            "wordPinyin": get(row, "WordPinyin"),
            "wordJa": get(row, "WordJa"),
            "sentenceZh": get(row, "SentenceZh"),
            "sentencePinyin": get(row, "SentencePinyin"),
            "sentenceJa": get(row, "SentenceJa"),
            "imageId": f"image_{no}",
            "imageFile": Path(image_path).name,
        }

        image = {
            "id": f"image_{no}",
            "cardId": card_id,
            "cardNo": no,
            "imageFile": Path(image_path).name,
            "imagePath": image_path,
            "imageStatus": "pending",
            "scene": get(row, "Scene"),
            "story": get(row, "Story"),
            "subject": get(row, "Subject"),
            "composition": get(row, "Composition"),
            "camera": get(row, "Camera"),
            "time": get(row, "Time"),
            "weather": get(row, "Weather"),
            "emotion": get(row, "Emotion"),
            "palette": get(row, "Palette"),
            "negativeSpace": get(row, "NegativeSpace"),
            "primaryMeaning": get(row, "Primary Meaning"),
            "secondaryMeaning": get(row, "Secondary Meaning"),
            "emotionSignal": get(row, "Emotion Signal"),
            "actionSignal": get(row, "Action Signal"),
            "environmentSignal": get(row, "Environment Signal"),
            "timeSignal": get(row, "Time Signal"),
            "visualPriority": [
                get(row, "Visual Priority 1"),
                get(row, "Visual Priority 2"),
                get(row, "Visual Priority 3"),
            ],
            "visualConcept": get(row, "Visual Concept"),
            "sceneV36": get(row, "Scene v3.6 Rebuilt"),
            "imagePromptV36": get(row, "Image Prompt v3.6"),
        }

        cards.append(card)
        images.append(image)

        if chapter_no and chapter_no not in chapters:
            chapters[chapter_no] = {"chapterNo": chapter_no, "chapterTitle": card["chapterTitle"]}
        if playlist_no and playlist_no not in playlists:
            playlists[playlist_no] = {"playlistNo": playlist_no, "playlistTitle": card["playlistTitle"], "chapterNo": chapter_no}

    OUT.mkdir(exist_ok=True)
    files = {
        "cards.json": cards,
        "images.json": images,
        "chapters.json": list(chapters.values()),
        "playlists.json": list(playlists.values()),
    }

    for filename, payload in files.items():
        (OUT / filename).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"created {len(cards)} cards / {len(images)} images")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("Usage: python tools/build_data_from_xlsx.py <xlsx>")
    main(sys.argv[1])
